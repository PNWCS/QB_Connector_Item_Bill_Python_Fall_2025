"""Excel reader for ItemBill data in company_data.xlsx.

This reader expects the worksheet to contain the following columns (exact
column headings, case-insensitive when trimmed):

- "Supplier Name"
- "Invoice Date"
- "Invoice Num"

It reads the "account debit vendor" worksheet, converts Excel date values to
Python `date` objects for `invoice_date`, and always returns `invoice_number` as a
string (even if the cell is numeric). Rows missing supplier or invoice number are
skipped. If present, the composite `id` is built from "Parent ID" and
"Child ID" columns as "<Parent ID>-<Child ID>".
"""

from __future__ import annotations

from pathlib import Path
import sys
from typing import List, Optional
from datetime import date, datetime

from openpyxl import load_workbook

from .models import ItemBill, Part


def _normalise(header: Optional[str]) -> str:
    return str(header).strip().lower() if header is not None else ""


temp_part = Part(name="Piston", quantity="10")


def extract_item_bills(workbook_path: Path) -> List[ItemBill]:
    """Read `workbook_path` and return a list of ItemBill objects.

    Raises FileNotFoundError when the workbook doesn't exist. Rows missing
    required fields are skipped.
    """

    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    try:
        sheet_name = "account debit vendor"
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Worksheet '{sheet_name}' not found in workbook")
        sheet = workbook[sheet_name]

        rows = sheet.iter_rows(values_only=True)
        headers_row = next(rows, None)
        if headers_row is None:
            workbook.close()
            return []

        headers = [_normalise(h) for h in headers_row]
        header_index = {h: idx for idx, h in enumerate(headers)}

        supplier_key = _normalise("Supplier Name")
        date_key = _normalise("Invoice Date")
        invoice_key = _normalise("Invoice Num")
        parent_id_key = _normalise("Parent ID")
        child_id_key = _normalise("Child ID")

        supplier_idx = header_index.get(supplier_key)
        date_idx = header_index.get(date_key)
        invoice_idx = header_index.get(invoice_key)
        parent_id_idx = header_index.get(parent_id_key)
        child_id_idx = header_index.get(child_id_key)

        bills: List[ItemBill] = []
        for row in rows:
            # supplier
            supplier_val = None
            if supplier_idx is not None and supplier_idx < len(row):
                supplier_val = row[supplier_idx]
            if supplier_val is None:
                continue
            supplier_name = str(supplier_val).strip()
            if not supplier_name:
                continue

            # invoice date
            invoice_date_val = None
            if date_idx is not None and date_idx < len(row):
                invoice_date_val = row[date_idx]

            # Normalize to a date object or None
            invoice_date: date | None
            if isinstance(invoice_date_val, datetime):
                invoice_date = invoice_date_val.date()
            elif isinstance(invoice_date_val, date):
                invoice_date = invoice_date_val
            elif invoice_date_val is None:
                invoice_date = None
            else:
                s = str(invoice_date_val).strip()
                if not s:
                    invoice_date = None
                else:
                    # Try ISO-8601 parsing
                    try:
                        invoice_date = datetime.fromisoformat(s).date()
                    except Exception:
                        invoice_date = None

            # invoice number
            invoice_number_val = None
            if invoice_idx is not None and invoice_idx < len(row):
                invoice_number_val = row[invoice_idx]
            if invoice_number_val is None:
                # skip if invoice number missing
                continue

            # Always normalise invoice number to string
            if isinstance(invoice_number_val, float):
                # Trim trailing .0 if it's an integer-like float
                if invoice_number_val.is_integer():
                    invoice_number = str(int(invoice_number_val))
                else:
                    invoice_number = ("%f" % invoice_number_val).rstrip("0").rstrip(".")
            else:
                invoice_number = str(invoice_number_val).strip()

            # Build composite id from Parent ID and Child ID if present
            parent_id_val = None
            if parent_id_idx is not None and parent_id_idx < len(row):
                parent_id_val = row[parent_id_idx]
            child_id_val = None
            if child_id_idx is not None and child_id_idx < len(row):
                child_id_val = row[child_id_idx]

            def _to_str(v):
                if v is None:
                    return ""
                try:
                    if isinstance(v, float) and v.is_integer():
                        return str(int(v))
                except Exception:
                    pass
                return str(v).strip()

            pid = _to_str(parent_id_val)
            cid = _to_str(child_id_val)
            composite_id = f"{pid}-{cid}" if pid or cid else None

            bills.append(
                ItemBill(
                    id=composite_id,
                    supplier_name=supplier_name,
                    invoice_date=invoice_date,
                    invoice_number=invoice_number,
                    parts=[temp_part],  # No parts info in Excel
                    source="excel",
                )
            )

    finally:
        workbook.close()

    return bills


__all__ = ["extract_item_bills"]


if __name__ == "__main__":  # pragma: no cover - manual execution helper
    if len(sys.argv) < 2:
        print(
            "Usage: python -m quickbook_vendor_item_bills.excel_reader <workbook_path>"
        )
        raise SystemExit(2)

    workbook_arg = sys.argv[1]
    try:
        bills = extract_item_bills(Path(workbook_arg))
    except Exception as e:  # simple CLI output for assignment demo
        print(f"Error reading workbook: {e}")
        raise SystemExit(1)

    for b in bills:
        print(str(b))
